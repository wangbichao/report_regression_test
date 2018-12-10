import io
import sys
import numpy
import openpyxl


# Change the default encoding of standard output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')

#ccl_report_file = "Z:\\bin_build\\ariel_build\\ariel_a0-x86_64-linux-dbg\\test_dcn_suite.txt"
#ccl_report_file = "Z:\\bin_build\\b0_ariel_build\\ariel_b0-x86_64-linux-dbg\\test_dcn_suite.txt"
#ccl_report_file = "X:\\bin_build\\bin_pubilc_ariel\\ariel_a0-x86_64-linux-dbg\\test_dcn_suite.txt"
ccl_report_file = "X:\\bin_build\\test_dcn_suite.txt"
status_tracker_file = "C:\\Users\\bichao\\Desktop\\Ariel_DCN2KA_IP_Diagnostics_Status_Tracker.xlsx"


# read the Status Tracker
wb = openpyxl.load_workbook(status_tracker_file)
all_sheets = wb.sheetnames
print("Get the status tracker all sheet ... ")
print(all_sheets)


source_summary = open(ccl_report_file, "r+", encoding='utf8')
suites_list = []
mismatch_list = []
total_suites_num = 0

temp_str = source_summary.readline()
while temp_str:
    suites_id = temp_str.split('.')[0]
    suites_list.append(suites_id)
    temp_str = source_summary.readline()

#print(suites_list)
source_summary.close()

count_suites_dict = dict(zip(*numpy.unique(suites_list, return_counts=True)))
print(count_suites_dict)
for i in count_suites_dict:
#    print(str(i) + ":" + str(count_suites_dict[i]))
    total_suites_num += count_suites_dict[i]
print("STF total varaiton :" + str(total_suites_num))


def find_false_in_sheet(sheet):
    for column in sheet.iter_cols():
        for cell in column:
            if cell.value in count_suites_dict.keys():
                if cell.column == 'F':
                    if sheet.cell(row=cell.row, column=29).value != count_suites_dict[cell.value]:
                        mismatch_list.append(sheet.cell(row=cell.row, column=3).value + "  " + sheet.cell(row=cell.row, column=6).value + " STF num: " + str(count_suites_dict[cell.value]) + " status tracker num: " + str(sheet.cell(row=cell.row, column=29).value))


for i in range(len(all_sheets)):
        sheet = wb[all_sheets[i]]
        find_false_in_sheet(sheet)

print("get mismatch suite....")
for i in mismatch_list:
    print(i)
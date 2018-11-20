import io
import sys
import openpyxl


# Change the default encoding of standard output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')

ccl_report_file = "C:\\Users\\bichao\Desktop\\Daily_report\\Ariel-CCL-summary.txt"
status_tracker_file = "C:\\Users\\bichao\\Desktop\\Ariel_DCN2KA_IP_Diagnostics_Status_Tracker.xlsx"
report_result_file = "C:\\Users\\bichao\Desktop\\Daily_report\\report_result.txt"

# read the CCL summary file
source_summary = open(ccl_report_file, "r+", encoding='utf8')
report_result = open(report_result_file, "w+", encoding='utf8')
hang_case_list = []
skip_case_list = []
fail_case_list = []
hang_case_list_detail = []
skip_case_list_detail = []
fail_case_list_detail = []
hang_case_list_detail_zip = []
skip_case_list_detail_zip = []
fail_case_list_detail_zip = []

# read the Status Tracker
wb = openpyxl.load_workbook(status_tracker_file)
all_sheets = wb.sheetnames
print("Get the status tracker all sheet ... ")
print(all_sheets)

# create the new report
ower_hang_case = []
ower_skip_case = []
ower_fail_case = []


def find_false_in_sheet(sheet):
    for column in sheet.iter_cols():
        for cell in column:
            if cell.value in hang_case_list:
#                print(cell.value + sheet.cell(row=cell.row, column=3).value)
                ower_hang_case.append(sheet.cell(row=cell.row, column=3).value + ' \t ' + cell.value + "(Hang)")
            if cell.value in skip_case_list:
#                print(cell.value + sheet.cell(row=cell.row, column=3).value)
                ower_skip_case.append(sheet.cell(row=cell.row, column=3).value + ' \t ' + cell.value + "(Skip)")
            if cell.value in fail_case_list:
#                print(cell.value + sheet.cell(row=cell.row, column=3).value)
                ower_fail_case.append(sheet.cell(row=cell.row, column=3).value + ' \t ' + cell.value + "(Fail)")


for temp_str in source_summary:
    if temp_str.find("Hang") != -1:
        hang_case_id_detail = temp_str.split()[0]
        if hang_case_id_detail not in hang_case_list_detail:
            hang_case_list_detail.append(hang_case_id_detail)
        hang_case_id = temp_str.split('.')[0]
        if hang_case_id not in hang_case_list:
            hang_case_list.append(hang_case_id)
    if temp_str.find("Skip") != -1:
        skip_case_id_detail = temp_str.split()[0]
        if skip_case_id_detail not in skip_case_list_detail:
            skip_case_list_detail.append(skip_case_id_detail)
        skip_case_id = temp_str.split('.')[0]
        if skip_case_id not in skip_case_list:
            skip_case_list.append(skip_case_id)
    if temp_str.find("Fail") != -1:
        fail_case_id_detail = temp_str.split()[0]
        if fail_case_id_detail not in fail_case_list_detail:
            fail_case_list_detail.append(fail_case_id_detail)
        fail_case_id = temp_str.split('.')[0]
        if fail_case_id not in fail_case_list:
            fail_case_list.append(fail_case_id)


print("Get the CCL summary all HANG case ... ")
print(len(hang_case_list))
print(hang_case_list)
print("Get the CCL summary all SKIP case ... ")
print(len(skip_case_list))
print(skip_case_list)
print("Get the CCL summary all FAIL case ... ")
print(len(fail_case_list))
print(fail_case_list)

for i in range(len(all_sheets)):
        sheet = wb[all_sheets[i]]
        find_false_in_sheet(sheet)

print("Report all HANG case with owner ... ")
ower_hang_case.sort()
print(len(ower_hang_case))
print(ower_hang_case)
report_result.write("All HANG case ...\n")
for report in ower_hang_case:
    report_result.write(report+"\n")

print("Report all SKIP case with owner ... ")
ower_skip_case.sort()
print(len(ower_skip_case))
print(ower_skip_case)
report_result.write("All SKIP case ...\n")
for report in ower_skip_case:
    report_result.write(report+"\n")

print("Report all FAIL case with owner ... ")
ower_fail_case.sort()
print(len(ower_fail_case))
print(ower_fail_case)
report_result.write("All FAIL case ...\n")
for report in ower_fail_case:
    report_result.write(report+"\n")

report_result.write("Get the CCL summary all HANG case ... \n")
report_result.write(str(len(hang_case_list_detail)) + "\n")
#for report in hang_case_list_detail:
#    report_result.write(report + "\n")
for i in range(len(hang_case_list_detail)):
#    print(i)
    if i == 0:
        new_temp_str = hang_case_list_detail[i]
    else:
        if hang_case_list_detail[i-1].split('.')[0] == hang_case_list_detail[i].split('.')[0]:
            if new_temp_str:
                new_temp_str = new_temp_str + "," + str(hang_case_list_detail[i].split('.')[1])
        else:
            hang_case_list_detail_zip.append(new_temp_str)
            new_temp_str = hang_case_list_detail[i]
        if i == (len(hang_case_list_detail) - 1):
            hang_case_list_detail_zip.append(new_temp_str)
#    print(new_temp_str)
#    print(hang_case_list_detail_zip)
for report in hang_case_list_detail:
    report_result.write(report + "\n")
report_result.write("Get the CCL summary all SKIP case ... \n ")
report_result.write(str(len(skip_case_list_detail)) + "\n")
#for report in skip_case_list_detail:
#    report_result.write(report + "\n")
for i in range(len(skip_case_list_detail)):
#    print(i)
    if i == 0:
        new_temp_str = skip_case_list_detail[i]
    else:
        if skip_case_list_detail[i-1].split('.')[0] == skip_case_list_detail[i].split('.')[0]:
            if new_temp_str:
                new_temp_str = new_temp_str + "," + str(skip_case_list_detail[i].split('.')[1])
        else:
            hang_case_list_detail_zip.append(new_temp_str)
            new_temp_str = skip_case_list_detail[i]
        if i == (len(skip_case_list_detail) - 1):
            hang_case_list_detail_zip.append(new_temp_str)
#    print(new_temp_str)
#    print(hang_case_list_detail_zip)
for report in hang_case_list_detail_zip:
    report_result.write(report + "\n")
report_result.write("Get the CCL summary all FAIL case ... \n ")
report_result.write(str(len(fail_case_list_detail)) + "\n")
#for report in fail_case_list_detail:
#    report_result.write(report + "\n")
for i in range(len(fail_case_list_detail)):
#    print(i)
    if i == 0:
        new_temp_str = fail_case_list_detail[i]
    else:
        if fail_case_list_detail[i-1].split('.')[0] == fail_case_list_detail[i].split('.')[0]:
            if new_temp_str:
                new_temp_str = new_temp_str + "," + str(fail_case_list_detail[i].split('.')[1])
        else:
            fail_case_list_detail_zip.append(new_temp_str)
            new_temp_str = fail_case_list_detail[i]
        if i == (len(fail_case_list_detail) - 1):
            fail_case_list_detail_zip.append(new_temp_str)
#    print(new_temp_str)
#    print(fail_case_list_detail_zip)
for report in fail_case_list_detail_zip:
    report_result.write(report + "\n")
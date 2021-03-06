import io
import sys
import openpyxl
import re


# Change the default encoding of standard output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')

status_tracker_file = "C:\\Users\\bichao\\Desktop\\Ariel_DCN2KA_IP_Diagnostics_Status_Tracker.xlsx"
report_result_file = "C:\\Users\\bichao\Desktop\\Daily_report\\get_Kandiah_case.txt"

report_result = open(report_result_file, "w+", encoding='utf8')


# read the Status Tracker
wb = openpyxl.load_workbook(status_tracker_file)
all_sheets = wb.sheetnames
print("Get the status tracker all sheet ... ")
print(all_sheets)

# report sprite and legacy case
kandiah_case = []


def find_false_in_sheet(sheet):
    for column in sheet.iter_cols():
        for cell in column:
            if re.search('Kandiah|kandiah', str(cell.value)):
#                print(cell.value + sheet.cell(row=cell.row, column=3).value + sheet.cell(row=cell.row, column=4).value)
                if cell.column == 'C':
                    kandiah_case.append(("NULL" if sheet.cell(row=cell.row, column=6).value == None else sheet.cell(row=cell.row, column=6).value)
                        + ' \t ' + ("NULL" if sheet.cell(row=cell.row, column=4).value == None else sheet.cell(row=cell.row, column=4).value)
                        + ' \t ' + ("NULL" if sheet.cell(row=cell.row, column=5).value == None else sheet.cell(row=cell.row, column=5).value)
                        + ' \t ' + cell.value)


for i in range(len(all_sheets)):
        sheet = wb[all_sheets[i]]
        find_false_in_sheet(sheet)


print("Get the status tracker all sheet with owner ... ")
kandiah_case.sort()
print(len(kandiah_case))
print(kandiah_case)
report_result.write("get sprite and legacy case ...\n")
for report in kandiah_case:
    report_result.write(report+"\n")
